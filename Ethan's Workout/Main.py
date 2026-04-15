"""Build Ethan's premium workout workbook.

Usage:
    python Main.py
    python Main.py --output Ethan_Elite_Workout_System.xlsx
"""

from __future__ import annotations

import argparse
from datetime import date, timedelta

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule, FormulaRule, IconSetRule
from openpyxl.styles import Alignment, Border, Font, GradientFill, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation

# =========================
# BRAND + VISUAL LANGUAGE
# =========================
BRAND_NAME = "ETHAN ELITE PERFORMANCE"
TAGLINE = "Discipline • Precision • Progress"

NAVY = "0F172A"
INK = "111827"
SLATE = "334155"
STEEL = "64748B"
CYAN = "06B6D4"
PURPLE = "7C3AED"
GREEN = "059669"
RED = "DC2626"
GOLD = "F59E0B"
WHITE = "FFFFFF"
LIGHT = "F8FAFC"
LIGHT_BLUE = "E0F2FE"
LIGHT_GREEN = "DCFCE7"
LIGHT_RED = "FEE2E2"

THIN = Side(style="thin", color="CBD5E1")
MED = Side(style="medium", color="64748B")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
MED_BORDER = Border(left=MED, right=MED, top=MED, bottom=MED)


# =========================
# COMMON STYLE HELPERS
# =========================
def configure_sheet(ws, freeze: str | None = None, zoom: int = 120) -> None:
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = zoom
    if freeze:
        ws.freeze_panes = freeze


def banner(ws, start_row: int = 1, start_col: int = 2, end_col: int = 12) -> None:
    for col in range(start_col, end_col + 1):
        c = ws.cell(row=start_row, column=col)
        c.fill = GradientFill(stop=(NAVY, PURPLE))
        c.border = MED_BORDER

    ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=end_col)
    c = ws.cell(row=start_row, column=start_col)
    c.value = f"{BRAND_NAME} | {TAGLINE}"
    c.font = Font(name="Calibri", size=13, bold=True, color=WHITE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[start_row].height = 24


def title(ws, cell: str, text: str, subtitle: str | None = None) -> None:
    ws[cell] = text
    ws[cell].font = Font(name="Calibri", size=22, bold=True, color=NAVY)
    ws[cell].alignment = Alignment(horizontal="left", vertical="center")

    if subtitle:
        r = ws[cell].row + 1
        c = ws[cell].column
        ws.cell(row=r, column=c, value=subtitle)
        ws.cell(row=r, column=c).font = Font(name="Calibri", size=10, italic=True, color=STEEL)


def section_header(ws, row: int, col_start: int, col_end: int, text: str) -> None:
    ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
    cell = ws.cell(row=row, column=col_start, value=text)
    cell.fill = PatternFill(fill_type="solid", fgColor=SLATE)
    cell.font = Font(name="Calibri", bold=True, color=WHITE)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    for col in range(col_start, col_end + 1):
        ws.cell(row=row, column=col).border = MED_BORDER


def header_row(ws, row: int, col_start: int, col_end: int) -> None:
    for col in range(col_start, col_end + 1):
        c = ws.cell(row=row, column=col)
        c.fill = PatternFill(fill_type="solid", fgColor=NAVY)
        c.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
        c.border = THIN_BORDER
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def data_cell(cell) -> None:
    cell.border = THIN_BORDER
    cell.font = Font(name="Calibri", size=10, color=INK)
    cell.alignment = Alignment(horizontal="center", vertical="center")


def zebra(ws, row_start: int, row_end: int, col_start: int, col_end: int) -> None:
    for r in range(row_start, row_end + 1):
        if r % 2 == 0:
            for c in range(col_start, col_end + 1):
                ws.cell(row=r, column=c).fill = PatternFill(fill_type="solid", fgColor=LIGHT)


def set_widths(ws, widths: dict[str, float]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def add_kpi_card(ws, row: int, label: str, formula: str, value_format: str | None = None) -> None:
    ws.cell(row=row, column=2, value=label)
    ws.cell(row=row, column=3, value=formula)

    left = ws.cell(row=row, column=2)
    right = ws.cell(row=row, column=3)

    left.fill = PatternFill(fill_type="solid", fgColor="E2E8F0")
    left.font = Font(name="Calibri", bold=True, color=SLATE)
    left.border = MED_BORDER
    left.alignment = Alignment(horizontal="left")

    right.fill = PatternFill(fill_type="solid", fgColor=WHITE)
    right.font = Font(name="Calibri", size=12, bold=True, color=NAVY)
    right.border = MED_BORDER
    right.alignment = Alignment(horizontal="center")

    if value_format:
        right.number_format = value_format


# =========================
# BUSINESS CONTENT
# =========================
EXERCISE_LIBRARY = [
    ("Chest", "Barbell Bench Press", "Barbell", "Compound", "Strength", "4x5", "120", "Retract scapula, full control."),
    ("Chest", "Incline Dumbbell Press", "Dumbbell", "Compound", "Hypertrophy", "4x8", "90", "Slow eccentric, drive up hard."),
    ("Chest", "Cable Fly", "Cable", "Isolation", "Hypertrophy", "3x12", "60", "Soft elbows, chest squeeze."),
    ("Back", "Deadlift", "Barbell", "Compound", "Strength", "5x3", "150", "Brace hard and push floor away."),
    ("Back", "Pull-Up", "Bodyweight", "Compound", "Strength", "4xAMRAP", "90", "Chest to bar, no kip."),
    ("Back", "Chest-Supported Row", "Machine", "Compound", "Hypertrophy", "4x10", "75", "Lead with elbows."),
    ("Legs", "Back Squat", "Barbell", "Compound", "Strength", "5x5", "150", "Knees track toes, stay braced."),
    ("Legs", "Romanian Deadlift", "Barbell", "Compound", "Hypertrophy", "4x8", "90", "Hip hinge, neutral spine."),
    ("Legs", "Walking Lunge", "Dumbbell", "Compound", "Hypertrophy", "3x12/side", "60", "Control each step."),
    ("Shoulders", "Overhead Press", "Barbell", "Compound", "Strength", "5x5", "120", "Stack ribs under bar."),
    ("Shoulders", "Lateral Raise", "Dumbbell", "Isolation", "Hypertrophy", "4x15", "45", "Lead with elbows."),
    ("Arms", "Barbell Curl", "Barbell", "Isolation", "Hypertrophy", "4x10", "60", "No torso swing."),
    ("Arms", "Cable Pushdown", "Cable", "Isolation", "Hypertrophy", "4x12", "60", "Lock elbows at sides."),
    ("Core", "Hanging Leg Raise", "Bodyweight", "Core", "Control", "4x10", "60", "Posterior pelvic tilt."),
    ("Core", "Pallof Press", "Cable", "Core", "Stability", "3x12/side", "45", "Resist rotation."),
    ("Conditioning", "Assault Bike Intervals", "Bike", "Conditioning", "Work Capacity", "10 rounds", "60", "Hard sprint, steady breathing."),
    ("Conditioning", "Sled Push", "Sled", "Conditioning", "Power", "6x20m", "90", "Forward body angle, drive knees."),
    ("Mobility", "90/90 Hip Flow", "Bodyweight", "Mobility", "Recovery", "3x60s", "30", "Slow transitions."),
    ("Mobility", "Ankle Dorsiflexion Drill", "Band", "Mobility", "Recovery", "3x12", "30", "Keep heel down."),
    ("Recovery", "Zone 2 Walk", "Bodyweight", "Recovery", "Aerobic", "30-45 min", "0", "Nasal breathing if possible."),
]

SESSION_TYPES = ["Upper", "Lower", "Push", "Pull", "Legs", "Conditioning", "Recovery", "Full Body"]
GOALS = ["Strength", "Hypertrophy", "Power", "Work Capacity", "Technique", "Recovery"]


# =========================
# SHEET BUILDERS
# =========================
def build_start_here(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "Start Here"
    configure_sheet(ws, zoom=125)
    banner(ws)
    title(ws, "B3", "Ethan Elite Workout System", subtitle=f"Generated on {date.today().strftime('%B %d, %Y')}")

    section_header(ws, 6, 2, 10, "Mission")
    ws.merge_cells("B7:J10")
    ws["B7"] = (
        "This workbook is your training operating system.\n\n"
        "Plan each week, execute every session, track recovery and nutrition daily, "
        "and review dashboard signals weekly."
    )
    ws["B7"].alignment = Alignment(wrap_text=True, vertical="top")
    ws["B7"].font = Font(name="Calibri", size=11, color=INK)
    ws["B7"].fill = PatternFill(fill_type="solid", fgColor=LIGHT)
    ws["B7"].border = MED_BORDER

    section_header(ws, 12, 2, 10, "Workflow")
    steps = [
        "1) Set body metrics in Body Metrics (starting/current/target weight).",
        "2) Build your weekly split in Program Builder.",
        "3) Log every workout in Daily Log (completion, duration, RPE, notes).",
        "4) Track recovery and nutrition every day.",
        "5) Review Dashboard each week and adjust load/volume as needed.",
    ]
    for i, step in enumerate(steps, start=13):
        ws.cell(row=i, column=2, value=step)
        ws.cell(row=i, column=2).alignment = Alignment(wrap_text=True)
        ws.cell(row=i, column=2).border = THIN_BORDER
        ws.cell(row=i, column=2).fill = PatternFill(fill_type="solid", fgColor=LIGHT if i % 2 == 0 else WHITE)
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=10)

    set_widths(ws, {"B": 32, "C": 14, "D": 14, "E": 14, "F": 14, "G": 14, "H": 14, "I": 14, "J": 14})


def build_dashboard(wb: Workbook) -> None:
    ws = wb.create_sheet("Dashboard")
    configure_sheet(ws)
    banner(ws)
    title(ws, "B3", "Performance Dashboard", subtitle="Live KPI overview")

    add_kpi_card(ws, 6, "Planned Sessions", "=COUNTIF('Program Builder'!$C$5:$C$500,\"<>\")")
    add_kpi_card(ws, 7, "Completed Sessions", "=COUNTIF('Daily Log'!$D$5:$D$2000,\"Yes\")")
    add_kpi_card(ws, 8, "Completion Rate", "=IFERROR(C7/C6,0)", "0.0%")
    add_kpi_card(ws, 9, "Avg Session Duration", "=IFERROR(AVERAGE('Daily Log'!$J$5:$J$2000),0)", "0")
    add_kpi_card(ws, 10, "Avg Session RPE", "=IFERROR(AVERAGE('Daily Log'!$K$5:$K$2000),0)", "0.0")
    add_kpi_card(ws, 11, "Current Weight", "='Body Metrics'!$C$6", "0.0")
    add_kpi_card(ws, 12, "Weight Change", "='Body Metrics'!$C$5-'Body Metrics'!$C$6", "0.0")

    section_header(ws, 6, 5, 8, "Recovery + Compliance")
    metrics = [
        ("Last Workout", "=LOOKUP(2,1/('Daily Log'!$B$5:$B$2000<>\"\"),'Daily Log'!$B$5:$B$2000)", "mmm d, yyyy"),
        (
            "This Week Sessions",
            "=COUNTIFS('Daily Log'!$B$5:$B$2000,\">="
            "&TODAY()-WEEKDAY(TODAY(),2)+1,'Daily Log'!$B$5:$B$2000,\"<="
            "&TODAY(),'Daily Log'!$D$5:$D$2000,\"Yes\")",
            "0",
        ),
        (
            "This Month Sessions",
            "=COUNTIFS('Daily Log'!$B$5:$B$2000,\">="
            "&EOMONTH(TODAY(),-1)+1,'Daily Log'!$B$5:$B$2000,\"<="
            "&EOMONTH(TODAY(),0),'Daily Log'!$D$5:$D$2000,\"Yes\")",
            "0",
        ),
        ("Avg Sleep", "=IFERROR(AVERAGE('Recovery Tracker'!$C$5:$C$2000),0)", "0.0"),
        ("Avg Steps", "=IFERROR(AVERAGE('Recovery Tracker'!$E$5:$E$2000),0)", "#,##0"),
        ("Avg Calories", "=IFERROR(AVERAGE('Nutrition'!$F$5:$F$2000),0)", "#,##0"),
    ]

    for i, (name, formula, fmt) in enumerate(metrics, start=7):
        ws.cell(row=i, column=5, value=name)
        ws.cell(row=i, column=6, value=formula)
        ws[f"E{i}"].fill = PatternFill(fill_type="solid", fgColor=LIGHT_BLUE)
        ws[f"E{i}"].font = Font(bold=True, color=SLATE)
        ws[f"E{i}"].border = MED_BORDER
        ws[f"F{i}"].border = MED_BORDER
        ws[f"F{i}"].alignment = Alignment(horizontal="center")
        ws[f"F{i}"].number_format = fmt

    # Completion mini chart
    ws["H6"] = "Metric"
    ws["I6"] = "Value"
    ws["H7"] = "Planned"
    ws["I7"] = "=C6"
    ws["H8"] = "Completed"
    ws["I8"] = "=C7"
    ws["H9"] = "Missed"
    ws["I9"] = "=MAX(C6-C7,0)"
    header_row(ws, 6, 8, 9)
    for r in range(7, 10):
        for c in range(8, 10):
            data_cell(ws.cell(row=r, column=c))

    chart = BarChart()
    chart.title = "Plan vs Completion"
    chart.height = 6
    chart.width = 9
    data = Reference(ws, min_col=9, min_row=6, max_row=9)
    cats = Reference(ws, min_col=8, min_row=7, max_row=9)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, "H11")

    set_widths(ws, {"B": 28, "C": 18, "D": 3, "E": 22, "F": 18, "G": 3, "H": 12, "I": 12})


def build_program_builder(wb: Workbook) -> None:
    ws = wb.create_sheet("Program Builder")
    configure_sheet(ws, freeze="B5")
    banner(ws)
    title(ws, "B3", "Program Builder", subtitle="Plan your training split and exercise targets")

    cols = [
        "Week",
        "Day",
        "Session Type",
        "Goal",
        "Exercise",
        "Sets",
        "Target Reps",
        "Target Load",
        "Rest (sec)",
        "Target RPE",
        "Notes",
    ]
    for i, name in enumerate(cols, start=2):
        ws.cell(row=4, column=i, value=name)
    header_row(ws, 4, 2, 12)

    days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
    for r in range(5, 505):
        ws.cell(row=r, column=2, value=((r - 5) // 7) + 1)
        ws.cell(row=r, column=3, value=days[(r - 5) % 7])
        for c in range(2, 13):
            data_cell(ws.cell(row=r, column=c))

    zebra(ws, 5, 504, 2, 12)

    session_dv = DataValidation(type="list", formula1=f'"{",".join(SESSION_TYPES)}"', allow_blank=True)
    goal_dv = DataValidation(type="list", formula1=f'"{",".join(GOALS)}"', allow_blank=True)
    ws.add_data_validation(session_dv)
    ws.add_data_validation(goal_dv)
    session_dv.add("D5:D504")
    goal_dv.add("E5:E504")

    for rng, lo, hi in [
        ("G5:G504", 1, 10),
        ("H5:H504", 1, 30),
        ("I5:I504", 0, 2000),
        ("J5:J504", 15, 300),
        ("K5:K504", 1, 10),
    ]:
        dv = DataValidation(type="whole", operator="between", formula1=lo, formula2=hi, allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(rng)

    ws.conditional_formatting.add("K5:K504", ColorScaleRule(start_type="num", start_value=1, start_color="BBF7D0", end_type="num", end_value=10, end_color="FECACA"))
    ws.conditional_formatting.add("G5:G504", DataBarRule(start_type="num", start_value=0, end_type="num", end_value=10, color=CYAN))

    ws.auto_filter.ref = "B4:L504"
    set_widths(ws, {"B": 8, "C": 12, "D": 14, "E": 14, "F": 34, "G": 8, "H": 10, "I": 10, "J": 10, "K": 10, "L": 32})


def build_exercise_library(wb: Workbook) -> None:
    ws = wb.create_sheet("Exercise Library")
    configure_sheet(ws, freeze="B5")
    banner(ws)
    title(ws, "B3", "Exercise Library", subtitle="Curated exercise database for program design")

    cols = ["Category", "Exercise", "Equipment", "Type", "Goal", "Default Prescription", "Rest", "Coaching Cue"]
    for i, name in enumerate(cols, start=2):
        ws.cell(row=4, column=i, value=name)
    header_row(ws, 4, 2, 9)

    for idx, row in enumerate(EXERCISE_LIBRARY, start=5):
        for c, value in enumerate(row, start=2):
            ws.cell(row=idx, column=c, value=value)
            data_cell(ws.cell(row=idx, column=c))

    zebra(ws, 5, 4 + len(EXERCISE_LIBRARY), 2, 9)
    ws.auto_filter.ref = f"B4:I{4 + len(EXERCISE_LIBRARY)}"
    set_widths(ws, {"B": 14, "C": 30, "D": 12, "E": 12, "F": 14, "G": 20, "H": 10, "I": 44})


def build_daily_log(wb: Workbook) -> None:
    ws = wb.create_sheet("Daily Log")
    configure_sheet(ws, freeze="B5")
    banner(ws)
    title(ws, "B3", "Daily Log", subtitle="Session-by-session execution data")

    cols = [
        "Date",
        "Day",
        "Completed",
        "Session Type",
        "Exercise",
        "Sets",
        "Reps",
        "Load",
        "Top Set",
        "Duration (min)",
        "Session RPE",
        "Mood",
        "Notes",
    ]
    for i, name in enumerate(cols, start=2):
        ws.cell(row=4, column=i, value=name)
    header_row(ws, 4, 2, 14)

    start = date.today() - timedelta(days=30)
    weekdays = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]

    for r in range(5, 2005):
        d = start + timedelta(days=r - 5)
        ws.cell(row=r, column=2, value=d.isoformat())
        ws.cell(row=r, column=3, value=weekdays[d.weekday()])
        for c in range(2, 15):
            data_cell(ws.cell(row=r, column=c))

    zebra(ws, 5, 2004, 2, 14)

    yes_no = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    types = DataValidation(type="list", formula1=f'"{",".join(SESSION_TYPES)}"', allow_blank=True)
    moods = DataValidation(type="list", formula1='"Locked In,Good,Neutral,Flat,Tired"', allow_blank=True)
    ws.add_data_validation(yes_no)
    ws.add_data_validation(types)
    ws.add_data_validation(moods)
    yes_no.add("D5:D2004")
    types.add("E5:E2004")
    moods.add("M5:M2004")

    for rng, lo, hi in [
        ("G5:G2004", 0, 20),
        ("H5:H2004", 0, 50),
        ("I5:I2004", 0, 3000),
        ("J5:J2004", 0, 5000),
        ("K5:K2004", 1, 300),
        ("L5:L2004", 1, 10),
    ]:
        dv = DataValidation(type="whole", operator="between", formula1=lo, formula2=hi, allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(rng)

    ws.conditional_formatting.add("B5:N2004", FormulaRule(formula=['$D5="Yes"'], fill=PatternFill(fill_type="solid", fgColor=LIGHT_GREEN), font=Font(color=GREEN)))
    ws.conditional_formatting.add("B5:N2004", FormulaRule(formula=['$D5="No"'], fill=PatternFill(fill_type="solid", fgColor=LIGHT_RED), font=Font(color=RED)))
    ws.conditional_formatting.add("L5:L2004", ColorScaleRule(start_type="num", start_value=1, start_color="BBF7D0", end_type="num", end_value=10, end_color="FCA5A5"))
    ws.conditional_formatting.add("K5:K2004", DataBarRule(start_type="num", start_value=0, end_type="num", end_value=180, color=CYAN))

    ws.auto_filter.ref = "B4:N2004"
    set_widths(ws, {"B": 12, "C": 11, "D": 12, "E": 14, "F": 30, "G": 8, "H": 8, "I": 10, "J": 10, "K": 12, "L": 10, "M": 14, "N": 36})


def build_body_metrics(wb: Workbook) -> None:
    ws = wb.create_sheet("Body Metrics")
    configure_sheet(ws, freeze="E6")
    banner(ws)
    title(ws, "B3", "Body Metrics", subtitle="Track body changes and trends")

    section_header(ws, 5, 2, 3, "Snapshot")
    labels = ["Starting Weight", "Current Weight", "Weight Change", "Target Weight", "To Goal"]
    for i, label in enumerate(labels, start=6):
        ws.cell(row=i, column=2, value=label)
        ws[f"B{i}"].fill = PatternFill(fill_type="solid", fgColor=LIGHT_BLUE)
        ws[f"B{i}"].font = Font(bold=True, color=SLATE)
        ws[f"B{i}"].border = MED_BORDER
        ws[f"C{i}"].border = MED_BORDER
        ws[f"C{i}"].alignment = Alignment(horizontal="center")
        ws[f"C{i}"].number_format = "0.0"

    ws["C8"] = "=IFERROR(C6-C7,0)"
    ws["C10"] = "=IFERROR(C7-C9,0)"

    headers = ["Date", "Weight", "Waist", "Chest", "Hip", "Body Fat %", "Resting HR"]
    for i, h in enumerate(headers, start=5):
        ws.cell(row=5, column=i, value=h)
    header_row(ws, 5, 5, 11)

    for r in range(6, 406):
        for c in range(5, 12):
            data_cell(ws.cell(row=r, column=c))

    zebra(ws, 6, 405, 5, 11)

    ws.conditional_formatting.add("F6:F405", DataBarRule(start_type="num", start_value=90, end_type="num", end_value=350, color=PURPLE))
    ws.conditional_formatting.add("J6:J405", ColorScaleRule(start_type="num", start_value=5, start_color="BBF7D0", end_type="num", end_value=35, end_color="FCA5A5"))

    chart = LineChart()
    chart.title = "Weight Trend"
    chart.height = 6
    chart.width = 10
    data = Reference(ws, min_col=6, min_row=5, max_row=405)
    cats = Reference(ws, min_col=5, min_row=6, max_row=405)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, "B13")

    ws.auto_filter.ref = "E5:K405"
    set_widths(ws, {"B": 20, "C": 12, "D": 3, "E": 12, "F": 10, "G": 10, "H": 10, "I": 10, "J": 12, "K": 12})


def build_recovery_tracker(wb: Workbook) -> None:
    ws = wb.create_sheet("Recovery Tracker")
    configure_sheet(ws, freeze="B5")
    banner(ws)
    title(ws, "B3", "Recovery Tracker", subtitle="Sleep, stress, soreness, hydration, steps")

    headers = ["Date", "Sleep (hrs)", "Sleep Quality", "Steps", "Stress", "Soreness", "Hydration (L)", "Recovery Score", "Notes"]
    for i, h in enumerate(headers, start=2):
        ws.cell(row=4, column=i, value=h)
    header_row(ws, 4, 2, 10)

    for r in range(5, 2005):
        for c in range(2, 11):
            data_cell(ws.cell(row=r, column=c))
        ws.cell(row=r, column=9, value=f"=IF(C{r}=\"\",\"\",ROUND((C{r}*3 + D{r}*2 + (11-F{r}) + (11-G{r}) + H{r}*2),1))")

    zebra(ws, 5, 2004, 2, 10)

    for rng, lo, hi in [
        ("C5:C2004", 0, 14),
        ("D5:D2004", 1, 10),
        ("E5:E2004", 0, 50000),
        ("F5:F2004", 1, 10),
        ("G5:G2004", 1, 10),
        ("H5:H2004", 0, 10),
    ]:
        dv = DataValidation(type="whole", operator="between", formula1=lo, formula2=hi, allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(rng)

    ws.conditional_formatting.add("I5:I2004", IconSetRule(icon_style="3TrafficLights1", type="num", values=[35, 55, 75]))
    ws.conditional_formatting.add("C5:C2004", ColorScaleRule(start_type="num", start_value=4, start_color="FECACA", end_type="num", end_value=9, end_color="BBF7D0"))

    ws.auto_filter.ref = "B4:J2004"
    set_widths(ws, {"B": 12, "C": 12, "D": 13, "E": 11, "F": 10, "G": 10, "H": 12, "I": 12, "J": 36})


def build_nutrition(wb: Workbook) -> None:
    ws = wb.create_sheet("Nutrition")
    configure_sheet(ws, freeze="B5")
    banner(ws)
    title(ws, "B3", "Nutrition", subtitle="Daily intake compliance")

    headers = ["Date", "Plan Followed", "Protein", "Carbs", "Fat", "Calories", "Fiber", "Water (L)", "Supplements", "Notes"]
    for i, h in enumerate(headers, start=2):
        ws.cell(row=4, column=i, value=h)
    header_row(ws, 4, 2, 11)

    for r in range(5, 2005):
        for c in range(2, 12):
            data_cell(ws.cell(row=r, column=c))

    zebra(ws, 5, 2004, 2, 11)

    yes_no = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    ws.add_data_validation(yes_no)
    yes_no.add("C5:C2004")

    for rng, lo, hi in [
        ("D5:D2004", 0, 400),
        ("E5:E2004", 0, 700),
        ("F5:F2004", 0, 250),
        ("G5:G2004", 0, 8000),
        ("H5:H2004", 0, 120),
        ("I5:I2004", 0, 10),
    ]:
        dv = DataValidation(type="whole", operator="between", formula1=lo, formula2=hi, allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(rng)

    ws.conditional_formatting.add("D5:D2004", DataBarRule(start_type="num", start_value=0, end_type="num", end_value=300, color=GREEN))
    ws.conditional_formatting.add("G5:G2004", ColorScaleRule(start_type="num", start_value=1200, start_color="BFDBFE", end_type="num", end_value=3500, end_color="1D4ED8"))

    ws.auto_filter.ref = "B4:K2004"
    set_widths(ws, {"B": 12, "C": 14, "D": 10, "E": 10, "F": 10, "G": 11, "H": 10, "I": 10, "J": 22, "K": 34})


def build_pr_tracker(wb: Workbook) -> None:
    ws = wb.create_sheet("PR Tracker")
    configure_sheet(ws, freeze="B5")
    banner(ws)
    title(ws, "B3", "PR Tracker", subtitle="Strength progression and estimated 1RM")

    headers = ["Date", "Lift", "Category", "Weight", "Reps", "Estimated 1RM", "Body Weight", "Relative Strength", "Notes"]
    for i, h in enumerate(headers, start=2):
        ws.cell(row=4, column=i, value=h)
    header_row(ws, 4, 2, 10)

    lifts = ["Back Squat", "Bench Press", "Deadlift", "Overhead Press", "Pull-Up", "Barbell Row"]

    for r in range(5, 1205):
        ws.cell(row=r, column=3, value=lifts[(r - 5) % len(lifts)])
        ws.cell(row=r, column=4, value="Compound")
        ws.cell(row=r, column=7, value=f"=IF(OR(E{r}=\"\",F{r}=\"\"),\"\",ROUND(E{r}*(1+F{r}/30),1))")
        ws.cell(row=r, column=9, value=f"=IF(OR(G{r}=\"\",H{r}=\"\"),\"\",ROUND(G{r}/H{r},2))")
        for c in range(2, 11):
            data_cell(ws.cell(row=r, column=c))

    zebra(ws, 5, 1204, 2, 10)

    lift_dv = DataValidation(type="list", formula1='"Back Squat,Bench Press,Deadlift,Overhead Press,Pull-Up,Barbell Row"', allow_blank=True)
    cat_dv = DataValidation(type="list", formula1='"Compound,Accessory,Conditioning"', allow_blank=True)
    ws.add_data_validation(lift_dv)
    ws.add_data_validation(cat_dv)
    lift_dv.add("C5:C1204")
    cat_dv.add("D5:D1204")

    ws.conditional_formatting.add("G5:G1204", DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1800, color=GOLD))
    ws.conditional_formatting.add("I5:I1204", ColorScaleRule(start_type="num", start_value=0.5, start_color="FECACA", end_type="num", end_value=3.0, end_color="BBF7D0"))

    ws.auto_filter.ref = "B4:J1204"
    set_widths(ws, {"B": 12, "C": 20, "D": 12, "E": 10, "F": 8, "G": 13, "H": 12, "I": 14, "J": 35})


def build_habit_tracker(wb: Workbook) -> None:
    ws = wb.create_sheet("Habit Tracker")
    configure_sheet(ws, freeze="B5")
    banner(ws)
    title(ws, "B3", "Habit Tracker", subtitle="Daily discipline score")

    headers = ["Date", "Train", "Nutrition", "Hydration", "Sleep", "Steps", "Mobility", "Journal", "No Alcohol", "Morning Walk", "Read", "Meditate", "Recovery Work", "Daily Score"]
    for i, h in enumerate(headers, start=2):
        ws.cell(row=4, column=i, value=h)
    header_row(ws, 4, 2, 15)

    for r in range(5, 1205):
        ws.cell(row=r, column=15, value=f"=SUM(C{r}:N{r})")
        for c in range(2, 16):
            data_cell(ws.cell(row=r, column=c))

    zebra(ws, 5, 1204, 2, 15)

    binary = DataValidation(type="list", formula1='"1,0"', allow_blank=True)
    ws.add_data_validation(binary)
    for col in "CDEFGHIJKLMN":
        binary.add(f"{col}5:{col}1204")

    ws.conditional_formatting.add("O5:O1204", IconSetRule(icon_style="5Arrows", type="num", values=[3, 6, 8, 10, 12]))
    ws.conditional_formatting.add("O5:O1204", DataBarRule(start_type="num", start_value=0, end_type="num", end_value=12, color=CYAN))

    ws.auto_filter.ref = "B4:O1204"
    set_widths(ws, {"B": 12, "C": 7, "D": 10, "E": 10, "F": 7, "G": 7, "H": 9, "I": 8, "J": 10, "K": 12, "L": 8, "M": 8, "N": 13, "O": 11})


def build_weekly_planner(wb: Workbook) -> None:
    ws = wb.create_sheet("Weekly Planner")
    configure_sheet(ws, zoom=130)
    banner(ws)
    title(ws, "B3", "Weekly Planner", subtitle="Tactical planning board")

    days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
    row = 6
    for day in days:
        section_header(ws, row, 2, 10, day)
        ws.merge_cells(start_row=row + 1, start_column=2, end_row=row + 2, end_column=6)
        ws.merge_cells(start_row=row + 1, start_column=7, end_row=row + 2, end_column=10)
        ws.cell(row=row + 1, column=2, value="Top Priorities")
        ws.cell(row=row + 1, column=7, value="Training Focus")
        for rr in (row + 1, row + 2):
            for cc in range(2, 11):
                ws.cell(row=rr, column=cc).border = THIN_BORDER
                ws.cell(row=rr, column=cc).alignment = Alignment(vertical="top", wrap_text=True)
        row += 3

    set_widths(ws, {"B": 20, "C": 20, "D": 20, "E": 20, "F": 20, "G": 20, "H": 20, "I": 20, "J": 20})


def build_instructions(wb: Workbook) -> None:
    ws = wb.create_sheet("Instructions")
    configure_sheet(ws)
    banner(ws)
    title(ws, "B3", "Instructions", subtitle="Run the system like a pro")

    blocks = [
        (
            "Setup",
            [
                "1) Enter body metrics first (starting/current/target weight).",
                "2) Build your split and exercise plan in Program Builder.",
                "3) Set realistic weekly target session count.",
            ],
        ),
        (
            "Daily",
            [
                "1) Log every workout session immediately after training.",
                "2) Log recovery and nutrition daily (no skipped days).",
                "3) Mark completion honestly (Yes/No).",
            ],
        ),
        (
            "Weekly Review",
            [
                "1) Review completion rate and weekly volume on Dashboard.",
                "2) Check PR Tracker for progression and plateaus.",
                "3) Adjust volume/intensity for next week in Program Builder.",
            ],
        ),
    ]

    r = 6
    for block, items in blocks:
        section_header(ws, r, 2, 11, block)
        r += 1
        for item in items:
            ws.cell(row=r, column=2, value=item)
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=11)
            ws.cell(row=r, column=2).border = THIN_BORDER
            ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True)
            ws.cell(row=r, column=2).fill = PatternFill(fill_type="solid", fgColor=LIGHT if r % 2 == 0 else WHITE)
            r += 1
        r += 1

    set_widths(ws, {"B": 24, "C": 24, "D": 24, "E": 24, "F": 24, "G": 24, "H": 24, "I": 24, "J": 24, "K": 24})


# =========================
# ENTRYPOINT
# =========================
def build_workbook(output_path: str) -> None:
    wb = Workbook()
    wb.properties.creator = "Ethan Elite Automation"
    wb.properties.title = "Ethan Elite Workout System"

    build_start_here(wb)
    build_dashboard(wb)
    build_program_builder(wb)
    build_exercise_library(wb)
    build_daily_log(wb)
    build_body_metrics(wb)
    build_recovery_tracker(wb)
    build_nutrition(wb)
    build_pr_tracker(wb)
    build_habit_tracker(wb)
    build_weekly_planner(wb)
    build_instructions(wb)

    wb.save(output_path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate Ethan's premium workout workbook.")
    parser.add_argument(
        "--output",
        default="Ethan_Elite_Workout_System.xlsx",
        help="Output filename (default: Ethan_Elite_Workout_System.xlsx)",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    build_workbook(args.output)
    print(f"Workbook created: {args.output}")


if __name__ == "__main__":
    main()
