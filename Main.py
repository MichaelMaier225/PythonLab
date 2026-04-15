"""Single-file generator for Ethan Elite Workout System.

Run:
    python Main.py --output Ethan_Elite_Workout_System.xlsx
"""

from __future__ import annotations

import argparse
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

TITLE = "Ethan Elite Workout System"
VERSION = "2.0.0"
DATA_START = 6
DATA_END = 500

# Brand tokens
NAVY = "0F172A"
BLUE = "2563EB"
EMERALD = "10B981"
AMBER = "F59E0B"
RED = "EF4444"
SLATE_100 = "F1F5F9"
SLATE_200 = "E2E8F0"
WHITE = "FFFFFF"

THIN = Side(style="thin", color=SLATE_200)
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")


# --------------------------- Core helpers ---------------------------
def apply_title(ws: Worksheet, title: str, subtitle: str) -> None:
    ws.merge_cells("A1:L1")
    ws["A1"] = title
    ws["A1"].font = Font(name="Calibri", size=18, bold=True, color=WHITE)
    ws["A1"].fill = PatternFill(fill_type="solid", fgColor=NAVY)
    ws["A1"].alignment = LEFT

    ws.merge_cells("A2:L2")
    ws["A2"] = subtitle
    ws["A2"].font = Font(name="Calibri", size=10, color=NAVY)
    ws["A2"].fill = PatternFill(fill_type="solid", fgColor=SLATE_100)
    ws["A2"].alignment = LEFT


def headers(ws: Worksheet, row: int, cols: Iterable[str]) -> None:
    for idx, label in enumerate(cols, start=1):
        c = ws.cell(row=row, column=idx, value=label)
        c.font = Font(name="Calibri", size=11, bold=True, color=WHITE)
        c.fill = PatternFill(fill_type="solid", fgColor=BLUE)
        c.alignment = CENTER
        c.border = BORDER


def style_grid(ws: Worksheet, min_row: int, max_row: int, max_col: int) -> None:
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.font = Font(name="Calibri", size=11, color=NAVY)
            cell.alignment = CENTER
            cell.border = BORDER


def set_cols(ws: Worksheet, widths: dict[str, float]) -> None:
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 20
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def add_list_validation(ws: Worksheet, rng: str, formula: str) -> None:
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    dv.error = "Choose from dropdown values."
    ws.add_data_validation(dv)
    dv.add(rng)


def add_whole_validation(ws: Worksheet, rng: str, lo: int, hi: int) -> None:
    dv = DataValidation(type="whole", operator="between", formula1=str(lo), formula2=str(hi), allow_blank=True)
    dv.error = f"Use integer between {lo} and {hi}."
    ws.add_data_validation(dv)
    dv.add(rng)


def add_decimal_validation(ws: Worksheet, rng: str, lo: float, hi: float) -> None:
    dv = DataValidation(type="decimal", operator="between", formula1=str(lo), formula2=str(hi), allow_blank=True)
    dv.error = f"Use value between {lo} and {hi}."
    ws.add_data_validation(dv)
    dv.add(rng)


def standard_sheet_finish(ws: Worksheet, freeze: str, filter_ref: str, print_ref: str) -> None:
    ws.freeze_panes = freeze
    ws.auto_filter.ref = filter_ref
    ws.print_area = print_ref
    ws.print_title_rows = "1:5"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.orientation = "landscape"


# --------------------------- Sheet builders ---------------------------
def build_lists(wb: Workbook) -> None:
    ws = wb.create_sheet("LISTS")
    ws.sheet_state = "hidden"

    data = {
        "A": ["Beginner", "Intermediate", "Advanced"],
        "B": ["Push", "Pull", "Legs", "Upper", "Lower", "Full Body", "Cardio"],
        "C": ["Squat", "Bench Press", "Deadlift", "Overhead Press", "Row", "Pull-up"],
        "D": ["Poor", "Fair", "Good", "Great", "Excellent"],
        "E": ["Cut", "Recomp", "Lean Bulk", "Performance", "Maintenance"],
        "F": [str(i) for i in range(1, 11)],
        "G": ["Yes", "No"],
    }

    for col, values in data.items():
        for row, value in enumerate(values, start=1):
            ws[f"{col}{row}"] = value


def build_start_here(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "START_HERE"
    apply_title(ws, TITLE, f"Version {VERSION} • Setup first, then follow workflow left-to-right.")
    set_cols(ws, {"A": 34, "B": 34, "C": 24, "D": 22, "E": 18})

    ws["A4"] = "Setup Checklist"
    ws["A4"].font = Font(size=13, bold=True, color=NAVY)
    for i, text in enumerate([
        "1) Fill athlete profile",
        "2) Set goals and level",
        "3) Build weekly plan",
        "4) Log workouts daily",
        "5) Review weekly dashboard",
    ], start=5):
        ws[f"A{i}"] = text

    ws["C4"] = "Athlete Profile"
    ws["C4"].font = Font(size=13, bold=True, color=NAVY)
    fields = [
        ("C5", "Name", "D5"),
        ("C6", "Start Date", "D6"),
        ("C7", "Age", "D7"),
        ("C8", "Training Level", "D8"),
        ("C9", "Primary Goal", "D9"),
        ("C10", "Units", "D10"),
    ]
    for label, txt, target in fields:
        ws[label] = txt
        ws[label].font = Font(bold=True)
        ws[target] = ""

    ws["D10"] = "lb"
    add_list_validation(ws, "D8", "=LISTS!$A$1:$A$3")
    add_list_validation(ws, "D9", "=LISTS!$E$1:$E$5")
    ws.freeze_panes = "A5"


def build_program_builder(wb: Workbook) -> None:
    ws = wb.create_sheet("PROGRAM_BUILDER")
    apply_title(ws, "Program Builder", "Define weekly plan, session design, and completion status.")
    set_cols(ws, {"A": 8, "B": 8, "C": 16, "D": 26, "E": 8, "F": 8, "G": 10, "H": 10, "I": 12})

    headers(ws, 5, ["Week", "Day", "Session", "Exercise", "Sets", "Reps", "Target RPE", "Rest(s)", "Done?"])
    style_grid(ws, DATA_START, DATA_END, 9)

    for r in range(DATA_START, DATA_END + 1):
        ws[f"A{r}"] = f'=IF(B{r}="","",INT((ROW()-6)/7)+1)'

    add_whole_validation(ws, f"B{DATA_START}:B{DATA_END}", 1, 7)
    add_list_validation(ws, f"C{DATA_START}:C{DATA_END}", "=LISTS!$B$1:$B$7")
    add_list_validation(ws, f"D{DATA_START}:D{DATA_END}", "=LISTS!$C$1:$C$6")
    add_whole_validation(ws, f"E{DATA_START}:E{DATA_END}", 1, 12)
    add_whole_validation(ws, f"F{DATA_START}:F{DATA_END}", 1, 30)
    add_whole_validation(ws, f"G{DATA_START}:G{DATA_END}", 1, 10)
    add_whole_validation(ws, f"H{DATA_START}:H{DATA_END}", 30, 600)
    add_list_validation(ws, f"I{DATA_START}:I{DATA_END}", "=LISTS!$G$1:$G$2")

    ws.conditional_formatting.add(
        f"A{DATA_START}:I{DATA_END}",
        FormulaRule(formula=["$I6=\"No\""], fill=PatternFill(fill_type="solid", fgColor="FEF3C7")),
    )

    standard_sheet_finish(ws, "A6", f"A5:I{DATA_END}", f"A1:I{DATA_END}")


def build_workout_log(wb: Workbook) -> None:
    ws = wb.create_sheet("WORKOUT_LOG")
    apply_title(ws, "Workout Log", "Log each session; formulas compute volume and estimated 1RM.")
    set_cols(ws, {"A": 12, "B": 8, "C": 22, "D": 8, "E": 8, "F": 10, "G": 8, "H": 12, "I": 12, "J": 24})

    headers(ws, 5, ["Date", "Day", "Exercise", "Sets", "Reps", "Load", "RPE", "Volume", "Est 1RM", "Notes"])
    style_grid(ws, DATA_START, DATA_END, 10)

    for r in range(DATA_START, DATA_END + 1):
        ws[f"H{r}"] = f'=IF(OR(F{r}="",E{r}="",D{r}=""),"",F{r}*E{r}*D{r})'
        ws[f"I{r}"] = f'=IF(OR(F{r}="",E{r}=""),"",ROUND(F{r}*(1+E{r}/30),1))'

    add_whole_validation(ws, f"B{DATA_START}:B{DATA_END}", 1, 7)
    add_list_validation(ws, f"C{DATA_START}:C{DATA_END}", "=LISTS!$C$1:$C$6")
    add_whole_validation(ws, f"D{DATA_START}:D{DATA_END}", 1, 12)
    add_whole_validation(ws, f"E{DATA_START}:E{DATA_END}", 1, 30)
    add_decimal_validation(ws, f"F{DATA_START}:F{DATA_END}", 0, 1500)
    add_whole_validation(ws, f"G{DATA_START}:G{DATA_END}", 1, 10)

    ws.conditional_formatting.add(
        f"A{DATA_START}:I{DATA_END}",
        FormulaRule(formula=["AND($I6>0,$I6>=LARGE($I$6:$I$500,5))"], fill=PatternFill(fill_type="solid", fgColor="DCFCE7")),
    )

    for r in range(DATA_START, DATA_END + 1):
        for c in range(1, 11):
            ws.cell(row=r, column=c).protection = Protection(locked=False)
        ws[f"H{r}"].protection = Protection(locked=True)
        ws[f"I{r}"].protection = Protection(locked=True)
    ws.protection.sheet = True

    standard_sheet_finish(ws, "A6", f"A5:J{DATA_END}", f"A1:J{DATA_END}")


def add_line_chart(ws: Worksheet, title: str, data_col: int, cat_col: int, start: int, end: int, anchor: str, y_title: str) -> None:
    chart = LineChart()
    chart.title = title
    chart.y_axis.title = y_title
    chart.x_axis.title = "Week"
    data = Reference(ws, min_col=data_col, min_row=start - 1, max_row=end)
    cats = Reference(ws, min_col=cat_col, min_row=start, max_row=end)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 6
    chart.width = 10
    ws.add_chart(chart, anchor)


def build_progression(wb: Workbook) -> None:
    ws = wb.create_sheet("PROGRESSION")
    apply_title(ws, "Progression", "Weekly trends in performance, workload, and adherence.")
    set_cols(ws, {"A": 8, "B": 14, "C": 14, "D": 16, "E": 16, "F": 14})

    headers(ws, 5, ["Week", "Avg Est 1RM", "Total Volume", "Planned", "Completed", "Adherence %"])
    style_grid(ws, 6, 57, 6)

    for r in range(6, 58):
        week = r - 5
        ws[f"A{r}"] = week
        ws[f"B{r}"] = f'=IFERROR(AVERAGE(WORKOUT_LOG!$I$6:$I$500),"")'
        ws[f"C{r}"] = f'=IFERROR(SUM(WORKOUT_LOG!$H$6:$H$500),"")'
        ws[f"D{r}"] = f'=COUNTIF(PROGRAM_BUILDER!$A$6:$A$500,A{r})'
        ws[f"E{r}"] = f'=COUNTIFS(PROGRAM_BUILDER!$A$6:$A$500,A{r},PROGRAM_BUILDER!$I$6:$I$500,"Yes")'
        ws[f"F{r}"] = f'=IF(D{r}=0,"",E{r}/D{r})'
        ws[f"F{r}"].number_format = "0%"

    add_line_chart(ws, "Average Estimated 1RM", 2, 1, 6, 57, "H5", "Weight")
    add_line_chart(ws, "Adherence", 6, 1, 6, 57, "H20", "%")

    standard_sheet_finish(ws, "A6", "A5:F57", "A1:L57")


def build_body_metrics(wb: Workbook) -> None:
    ws = wb.create_sheet("BODY_METRICS")
    apply_title(ws, "Body Metrics", "Monitor bodyweight and measurements.")
    set_cols(ws, {"A": 12, "B": 12, "C": 12, "D": 12, "E": 12, "F": 14})

    headers(ws, 5, ["Date", "Bodyweight", "Waist", "Chest", "Body Fat %", "7d Avg BW"])
    style_grid(ws, 6, 250, 6)

    for r in range(6, 251):
        ws[f"F{r}"] = f'=IF(COUNT($B$6:B{r})<7,"",AVERAGE(INDEX($B:$B,ROW()-6):B{r}))'

    add_decimal_validation(ws, "B6:B250", 50, 500)
    add_decimal_validation(ws, "C6:D250", 10, 100)
    add_decimal_validation(ws, "E6:E250", 2, 60)

    chart = LineChart()
    chart.title = "Bodyweight"
    chart.y_axis.title = "Weight"
    data = Reference(ws, min_col=2, min_row=5, max_row=250)
    cats = Reference(ws, min_col=1, min_row=6, max_row=250)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 6
    chart.width = 10
    ws.add_chart(chart, "H5")

    standard_sheet_finish(ws, "A6", "A5:F250", "A1:L250")


def build_recovery_sleep(wb: Workbook) -> None:
    ws = wb.create_sheet("RECOVERY_SLEEP")
    apply_title(ws, "Recovery & Sleep", "Track readiness markers to guide training decisions.")
    set_cols(ws, {"A": 12, "B": 12, "C": 12, "D": 10, "E": 10, "F": 12, "G": 16})

    headers(ws, 5, ["Date", "Sleep Hours", "Quality", "Stress", "Soreness", "Readiness", "Adjustment Notes"])
    style_grid(ws, 6, 250, 7)

    for r in range(6, 251):
        ws[f"F{r}"] = f'=IF(COUNTA(B{r}:E{r})<4,"",ROUND(((B{r}/8)*4)+((MATCH(C{r},LISTS!$D$1:$D$5,0)/5)*2)+((11-D{r})/10*2)+((11-E{r})/10*2),1))'

    add_decimal_validation(ws, "B6:B250", 0, 14)
    add_list_validation(ws, "C6:C250", "=LISTS!$D$1:$D$5")
    add_whole_validation(ws, "D6:E250", 1, 10)

    ws.conditional_formatting.add(
        "A6:G250",
        FormulaRule(formula=["$F6<6"], fill=PatternFill(fill_type="solid", fgColor="FEE2E2")),
    )
    standard_sheet_finish(ws, "A6", "A5:G250", "A1:L250")


def build_nutrition(wb: Workbook) -> None:
    ws = wb.create_sheet("NUTRITION")
    apply_title(ws, "Nutrition", "Log macro intake and calorie compliance.")
    set_cols(ws, {"A": 12, "B": 12, "C": 10, "D": 10, "E": 10, "F": 10, "G": 16, "I": 14})

    headers(ws, 5, ["Date", "Calories", "Protein", "Carbs", "Fat", "Water (L)", "Calorie Compliance %"])
    style_grid(ws, 6, 250, 7)
    ws["I5"] = "Calorie Target"
    ws["I6"] = 2500

    for r in range(6, 251):
        ws[f"G{r}"] = f'=IF(OR(B{r}="",$I$6=""),"",1-ABS(B{r}-$I$6)/$I$6)'
        ws[f"G{r}"].number_format = "0%"

    add_decimal_validation(ws, "B6:B250", 800, 7000)
    add_decimal_validation(ws, "C6:E250", 0, 1000)
    add_decimal_validation(ws, "F6:F250", 0, 10)

    ws.conditional_formatting.add(
        "G6:G250",
        ColorScaleRule(
            start_type="num", start_value=0.6, start_color="FEE2E2",
            mid_type="num", mid_value=0.8, mid_color="FEF3C7",
            end_type="num", end_value=1.0, end_color="DCFCE7",
        ),
    )
    standard_sheet_finish(ws, "A6", "A5:G250", "A1:L250")


def build_weekly_review(wb: Workbook) -> None:
    ws = wb.create_sheet("WEEKLY_REVIEW")
    apply_title(ws, "Weekly Review", "Executive dashboard for outcomes and next actions.")
    set_cols(ws, {"A": 24, "B": 16, "C": 24, "D": 16, "E": 24, "F": 16, "G": 24, "H": 16})

    tiles = [
        ("A5", "Adherence %", '=IFERROR(PROGRESSION!F57,"")', "0%"),
        ("C5", "Avg Est 1RM", '=IFERROR(PROGRESSION!B57,"")', "0.0"),
        ("E5", "7d Avg BW", '=IFERROR(BODY_METRICS!F250,"")', "0.0"),
        ("G5", "Recovery Score", '=IFERROR(AVERAGE(RECOVERY_SLEEP!F244:F250),"")', "0.0"),
    ]
    for tcell, label, formula, fmt in tiles:
        vcell = f"{tcell[0]}6"
        ws[tcell] = label
        ws[tcell].font = Font(bold=True, color=WHITE)
        ws[tcell].fill = PatternFill(fill_type="solid", fgColor=BLUE)
        ws[vcell] = formula
        ws[vcell].font = Font(size=14, bold=True, color=NAVY)
        ws[vcell].fill = PatternFill(fill_type="solid", fgColor="DBEAFE")
        ws[vcell].number_format = fmt

    ws["A10"] = "Weekly Coach Notes"
    ws["A10"].font = Font(size=12, bold=True, color=NAVY)
    ws.merge_cells("A11:H22")
    ws["A11"] = "- Wins\n- Constraints\n- Adjustments for next week"
    ws["A11"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws["A11"].fill = PatternFill(fill_type="solid", fgColor=SLATE_100)
    ws.freeze_panes = "A5"


def build_help(wb: Workbook) -> None:
    ws = wb.create_sheet("HELP")
    apply_title(ws, "Instructions", "Simple operating guide for daily use.")
    set_cols(ws, {"A": 110})

    content = {
        "Workflow": [
            "1) Fill START_HERE.",
            "2) Set weekly plan in PROGRAM_BUILDER.",
            "3) Log every session in WORKOUT_LOG.",
            "4) Enter body/recovery/nutrition daily.",
            "5) Review WEEKLY_REVIEW weekly and update plan.",
        ],
        "Best Practices": [
            "Use dropdown lists to preserve data integrity.",
            "Do not edit formula cells in protected columns.",
            "Focus on 2-4 week trends, not day-to-day noise.",
        ],
        "Troubleshooting": [
            "If formulas look blank, check required inputs in the same row.",
            "If dropdowns are missing, keep LISTS sheet intact.",
            "If print output is odd, use Landscape and Fit to Width=1.",
        ],
    }

    row = 4
    for section, bullets in content.items():
        ws[f"A{row}"] = section
        ws[f"A{row}"].font = Font(size=13, bold=True, color=NAVY)
        row += 1
        for b in bullets:
            ws[f"A{row}"] = f"• {b}"
            row += 1
        row += 1

    ws.freeze_panes = "A4"


# --------------------------- Build orchestration ---------------------------
def build_workbook(output: Path) -> Path:
    wb = Workbook()
    build_start_here(wb)
    build_program_builder(wb)
    build_workout_log(wb)
    build_progression(wb)
    build_body_metrics(wb)
    build_recovery_sleep(wb)
    build_nutrition(wb)
    build_weekly_review(wb)
    build_help(wb)
    build_lists(wb)

    order = [
        "START_HERE",
        "PROGRAM_BUILDER",
        "WORKOUT_LOG",
        "PROGRESSION",
        "BODY_METRICS",
        "RECOVERY_SLEEP",
        "NUTRITION",
        "WEEKLY_REVIEW",
        "HELP",
        "LISTS",
    ]
    index = {name: i for i, name in enumerate(order)}
    wb._sheets.sort(key=lambda ws: index.get(ws.title, 999))

    wb.properties.title = TITLE
    wb.properties.creator = "Ethan Elite Single-File Generator"

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    return output


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Generate Ethan Elite Workout workbook.")
    p.add_argument("--output", type=Path, default=Path("Ethan_Elite_Workout_System.xlsx"))
    return p.parse_args()


def main() -> None:
    args = parse_args()
    path = build_workbook(args.output)
    print(f"Workbook generated: {path}")


if __name__ == "__main__":
    main()
